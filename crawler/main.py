import time

from crawler.functional.crawler import Crawler
import openpyxl
import threading


def open_file(path):
    list_crawler = []
    with open(path, 'r') as f:
        for idData in f:
            if "\n" in idData:
                id_url = idData.rstrip('\n')
            else:
                id_url = idData
            print(id_url)
            crawler = Crawler("https://amis.misa.vn/" + id_url)
            list_crawler.append(crawler)
    return list_crawler


def export_excel(path, i):
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    list_headers = ["STT", "Chuyên mục", "Tác giả", "URL cuối", "Số lượng từ", "Danh sách Internal Link trong bài",
                    "Danh sách External Link trong bài", "Tag", "Index/NoIndex", "Follow/Nofollow", "Ahref-Lang",
                    "Meta Title", "Meta Description", "Meta keyword", "Canonical", "Date Published", "Date Modified"]
    list_crawl = open_file(path)
    time.sleep(0.1)
    index = 1
    for header in list_headers:
        column_header = my_sheet.cell(row=1, column=list_headers.index(header) + 1)
        column_header.value = header

    for idCrawler in list_crawl:
        if idCrawler.get_meta_description() == "404 - MISA AMIS":
            print("Can't crawl Data")
        else:
            index += 1
            # Tạo cột số thứ tự
            data = my_sheet.cell(row=index, column=1)
            data.value = index - 1
            # Tạo cột category
            data = my_sheet.cell(row=index, column=2)
            data.value = idCrawler.list_to_string(idCrawler.get_list_category())
            # Tạo cột author
            author_data = my_sheet.cell(row=index, column=3)
            author_data.value = idCrawler.get_author()
            # Tạo cột url
            url_data = my_sheet.cell(row=index, column=4)
            url_data.value = idCrawler.get_url()

            # Tạo cột số lượng dữ liệu
            number_of_data = my_sheet.cell(row=index, column=5)
            number_of_data.value = idCrawler.get_num_word()

            # Tạo cột internalUrl

            internal_data = my_sheet.cell(row=index, column=6)
            internal_data.value = idCrawler.list_to_string(idCrawler.get_internal_link_list())

            # Tạo cột externalUrl

            external_data = my_sheet.cell(row=index, column=7)
            external_data.value = idCrawler.list_to_string(idCrawler.get_external_link_list())

            # Tạo cột tag
            tag_list_data = my_sheet.cell(row=index, column=8)
            tag_list_data.value = ','.join(str(tagData) for tagData in idCrawler.get_tag_list())

            # Tạo cột xác định index hay noindex
            is_index_data = my_sheet.cell(row=index, column=9)
            is_index_data.value = idCrawler.index_or_no()

            # Tạo cột xác định folow hay noFollow
            is_follow_data = my_sheet.cell(row=index, column=10)
            is_follow_data.value = idCrawler.follow_or_no()

            # Tạo cột hrefLanguage
            href_lang_data = my_sheet.cell(row=index, column=11)
            href_lang_data.value = idCrawler.get_hreflang()

            # Tạo cột meta title
            meta_title_data = my_sheet.cell(row=index, column=12)
            meta_title_data.value = str(idCrawler.get_meta_title())

            # Tạo cột meta description
            meta_description_data = my_sheet.cell(row=index, column=13)
            meta_description_data.value = idCrawler.get_meta_description()

            # Tạo cột canonicalUrl
            meta_canonical_data = my_sheet.cell(row=index, column=15)
            meta_canonical_data.value = idCrawler.get_canonical_link()

            # Tạo cột date published
            date_publish_data = my_sheet.cell(row=index, column=16)
            date_publish_data.value = idCrawler.get_date_published()

            # Tạo cọt date modified
            date_modified_data = my_sheet.cell(row=index, column=17)
            date_modified_data.value = idCrawler.get_date_modified()
            print(index - 1)
            my_wb.save("crawl_data" + i + ".xlsx")


# main function
if __name__ == '__main__':
    print("start export")
    t1 = threading.Thread(target=export_excel, args=("idWebsite1.txt", "1",))
    t1.start()
    t2 = threading.Thread(target=export_excel, args=("idWebsite2.txt", "2",))
    t2.start()
    t3 = threading.Thread(target=export_excel, args=("idWebsite3.txt", "3",))
    t3.start()
    t4 = threading.Thread(target=export_excel, args=("idWebsite4.txt", "4",))
    t4.start()
    t5 = threading.Thread(target=export_excel, args=("idWebsite5.txt", "5",))
    t5.start()
    t6 = threading.Thread(target=export_excel, args=("idWebsite6.txt", "6",))
    t6.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()

    print("end export")
